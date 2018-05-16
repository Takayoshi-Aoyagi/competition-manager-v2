# -*- coding: utf-8 -*-

import os

from openpyxl import load_workbook

def uni(row, index):
    cell = row[index]
    if cell is None or cell.value is None:
        return None
    return unicode(cell.value)

class ExcelInfile:

    @staticmethod
    def get_column_defs():
        columns = (
            ("番号", 0, "number"),
            ("出場選手", 1, "name"),
            ("性別", 3, "gender"),
            ("段、級", 4, "grade"),
            ("所属", 5, "dojo"),
            ("トゥル", 6, "tul"),
            ("マッソギ", 9, "massogi"),
            ("スペシャル", 13, "special"),
            ("氏名ふりがな", 17, "kana")
        )
        return columns

class ExcelInfileReader:

    def __init__(self, file_path):
        self.file_path = file_path
        self.fname = os.path.basename(file_path)
        self.wb = load_workbook(filename=self.file_path)

    def is_valid_entrant(self, entrant):
        # None check
        for key in ['name', 'tul', 'massogi']:
            if entrant[key] is None:
                return False
        # Name check
        for name in [u'出場選手', u'審判員', u'川口　太郎', u'わらび　花子']:
            if entrant['name'] == name:
                return False
        return True
            
    def _read_by_row(self, row, column_defs):
        rowlen = column_defs[-1][1]
        if rowlen >= len(row):
            return None
        entrant = {}
        for column_def in column_defs:
            _, index, key = column_def
            entrant[key] = uni(row, index)
            entrant["fname"] = self.fname

        if self.is_valid_entrant(entrant) is False:
            return None
        return entrant
        
    def _read_by_sheet(self, sheet):
        entrants = []
        column_defs = ExcelInfile.get_column_defs()
        for row in sheet.iter_rows():
            entrant = self._read_by_row(row, column_defs)
            if entrant is None:
                continue
            entrants.append(entrant)
        return entrants
            
    def read(self):
        entrants = []
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            _entrants = self._read_by_sheet(sheet)
            entrants.extend(_entrants)
        return entrants
            
