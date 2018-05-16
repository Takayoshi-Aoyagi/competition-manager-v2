# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border

class ExcelWriter:

    def __init__(self, file_path):
        self.header_style = PatternFill(start_color='AA7fc1ff',
                                        end_color='AA7fc1ff',
                                        fill_type='solid')
        self.file_path = file_path
        self.wb = openpyxl.Workbook()
        for sheet in self.wb.worksheets:
            self.wb.remove(sheet)

    def _create_row(self, sheet, row_index, row, style=None):
        col_indexes = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        for i, col in enumerate(row):
            col_index = col_indexes[i]
            addr = "{0}{1}".format(col_index, row_index)
            sheet[addr].value = col
            if style is None:
                continue
            sheet[addr].fill = style

    def _create_sheet(self, sheet_name, rows):
        sheet = self.wb.create_sheet(sheet_name)
        for i, row in enumerate(rows):
            row_index = i + 1
            style = None
            if i == 0:
                style = self.header_style
            self._create_row(sheet, row_index, row, style=style)
        
    def write(self, data):
        for (sheet_name, sheet_data) in data.items():
            self._create_sheet(sheet_name, sheet_data)
        self.wb.save(self.file_path)
