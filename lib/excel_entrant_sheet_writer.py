# -*- coding: utf-8 -*-

import openpyxl

class ExcelEntrantSheetWriter:

    def __init__(self, entrants):
        self.entrants = entrants
        self.wb = openpyxl.Workbook()
        s0 = self.wb['Sheet']
        self.wb.remove(s0)

    def _get_row_data(self, entrant):
        keys = ("name", "grade", "dojo", "tul", "massogi", "special", "kana", "fname")
        row = []
        for key in keys:
            row.append(entrant[key])
        return row

    def _write_to_row(self, sheet, row_index, row):
        col_indexes = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        for i, col in enumerate(row):
            col_index = col_indexes[i]
            addr = "{0}{1}".format(col_index, row_index)
            sheet[addr].value = col
        
    def _write_to_sheet(self, sheet, rows):
        for i, row in enumerate(rows):
            row_index = i + 1
            self._write_to_row(sheet, row_index, row)
        
    def _create_all_entrants_sheets(self):
        sheet = self.wb.create_sheet(u'参加者一覧')
        rows = []
        for entrant in self.entrants:
            row = self._get_row_data(entrant)
            rows.append(row)
        self._write_to_sheet(sheet, rows)
            
        
    def create_sheets(self):
        self._create_all_entrants_sheets()
        
    def write(self, file_path):
        self.create_sheets()
        self.wb.save(file_path)
        
